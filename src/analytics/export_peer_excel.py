import sqlite3
import pandas as pd

from statistics import median

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.screener.engine import load_master_dataframe


DB_PATH = "db/nifty100.db"


# ==========================================================
# COLORS
# ==========================================================

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="90EE90",
    end_color="90EE90"
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFF59D",
    end_color="FFF59D"
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FF9999",
    end_color="FF9999"
)

GOLD_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD700",
    end_color="FFD700"
)


# ==========================================================
# DATABASE
# ==========================================================

def get_connection():
    return sqlite3.connect(DB_PATH)


# ==========================================================
# LOAD TABLES
# ==========================================================

def load_peer_percentiles():

    conn = get_connection()

    df = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn
    )

    conn.close()

    return df


def load_peer_groups():

    conn = get_connection()

    df = pd.read_sql(
        "SELECT * FROM peer_groups",
        conn
    )

    conn.close()

    return df


# ==========================================================
# LOAD COMPANY DATA
# ==========================================================

def load_company_data():

    df = load_master_dataframe()

    df = df[
        df["year"] == "Mar 2024"
    ]

    return df


# ==========================================================
# MAIN
# ==========================================================

if __name__ == "__main__":

    peer_percentiles = load_peer_percentiles()
    peer_groups = load_peer_groups()
    companies = load_company_data()

    workbook = Workbook()

    workbook.remove(workbook.active)

    groups = sorted(
        peer_groups["peer_group_name"].unique()
    )

    print(
        f"Creating {len(groups)} sheets...\n"
    )

    for group in groups:

        print(f"Creating sheet: {group}")

        sheet = workbook.create_sheet(
            title=group
        )

        group_companies = peer_groups[
            peer_groups["peer_group_name"] == group
        ]

        merged = group_companies.merge(
            companies,
            on="company_id",
            how="left"
        )

        group_percentiles = peer_percentiles[
            peer_percentiles["peer_group_name"] == group
        ]

        percentile_lookup = (
            group_percentiles
            .set_index(
                ["company_id", "metric"]
            )["percentile_rank"]
            .to_dict()
        )

        benchmark_company = group_companies.loc[
            group_companies["is_benchmark"] == 1,
            "company_id"
        ].iloc[0]

        # ==========================================================
        # COLUMN HEADERS
        # ==========================================================

        columns = [

            "company_id",
            "company_name",
            "composite_quality_score",

            "return_on_equity_pct",
            "ROE Percentile",

            "roce_percentage",
            "ROCE Percentile",

            "net_profit_margin_pct",
            "NPM Percentile",

            "debt_to_equity",
            "D/E Percentile",

            "free_cash_flow_cr",
            "FCF Percentile",

            "pat_cagr_5yr",
            "PAT CAGR Percentile",

            "revenue_cagr_5yr",
            "Revenue CAGR Percentile",

            "eps_cagr_5yr",
            "EPS CAGR Percentile",

            "interest_coverage",
            "ICR Percentile",

            "asset_turnover",
            "Asset Turnover Percentile"

        ]

        sheet.append(columns)


        # ==========================================================
        # WRITE COMPANY DATA
        # ==========================================================

        for _, row in merged.iterrows():

            sheet.append([

                row["company_id"],
                row["company_name"],
                row["composite_quality_score"],

                row["return_on_equity_pct"],
                percentile_lookup.get(
                    (row["company_id"], "return_on_equity_pct"),
                    ""
                ),

                row["roce_percentage"],
                percentile_lookup.get(
                    (row["company_id"], "roce_percentage"),
                    ""
                ),

                row["net_profit_margin_pct"],
                percentile_lookup.get(
                    (row["company_id"], "net_profit_margin_pct"),
                    ""
                ),

                row["debt_to_equity"],
                percentile_lookup.get(
                    (row["company_id"], "debt_to_equity"),
                    ""
                ),

                row["free_cash_flow_cr"],
                percentile_lookup.get(
                    (row["company_id"], "free_cash_flow_cr"),
                    ""
                ),

                row["pat_cagr_5yr"],
                percentile_lookup.get(
                    (row["company_id"], "pat_cagr_5yr"),
                    ""
                ),

                row["revenue_cagr_5yr"],
                percentile_lookup.get(
                    (row["company_id"], "revenue_cagr_5yr"),
                    ""
                ),

                row["eps_cagr_5yr"],
                percentile_lookup.get(
                    (row["company_id"], "eps_cagr_5yr"),
                    ""
                ),

                row["interest_coverage"],
                percentile_lookup.get(
                    (row["company_id"], "interest_coverage"),
                    ""
                ),

                row["asset_turnover"],
                percentile_lookup.get(
                    (row["company_id"], "asset_turnover"),
                    ""
                )

            ])

        # ======================================================
        # COLOR PERCENTILE COLUMNS
        # ======================================================

        percentile_columns = [
            5,   # ROE Percentile
            7,   # ROCE Percentile
            9,   # NPM Percentile
            11,  # D/E Percentile
            13,  # FCF Percentile
            15,  # PAT CAGR Percentile
            17,  # Revenue CAGR Percentile
            19,  # EPS CAGR Percentile
            21,  # ICR Percentile
            23   # Asset Turnover Percentile
        ]

        for r in range(2, sheet.max_row + 1):

            for c in percentile_columns:

                cell = sheet.cell(row=r, column=c)

                if isinstance(cell.value, (int, float)):

                    if cell.value >= 75:
                        cell.fill = GREEN_FILL

                    elif cell.value <= 25:
                        cell.fill = RED_FILL

                    else:
                        cell.fill = YELLOW_FILL

        # ======================================================
        # HIGHLIGHT BENCHMARK COMPANY
        # ======================================================

        for r in range(2, sheet.max_row + 1):

            if sheet.cell(r, 1).value == benchmark_company:

                for c in range(1, sheet.max_column + 1):

                    sheet.cell(r, c).fill = GOLD_FILL

        # ======================================================
        # PEER MEDIAN ROW
        # ======================================================

        median_row = [
            "Peer Median",
            ""
        ]

        for col in range(3, sheet.max_column + 1):

            values = []

            for r in range(2, sheet.max_row + 1):

                value = sheet.cell(r, col).value

                if isinstance(value, (int, float)):
                    values.append(value)

            if values:
                median_row.append(round(median(values), 2))
            else:
                median_row.append("")

        sheet.append(median_row)

    # ======================================================
    # SAVE WORKBOOK
    # ======================================================

    workbook.save("output/peer_comparison.xlsx")

    print("\nWorkbook saved successfully.")
    