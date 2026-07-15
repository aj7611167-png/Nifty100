import sqlite3
import yaml
import pandas as pd

DB_PATH = "db/nifty100.db"
CONFIG_PATH = "config/screener_config.yaml"


# ==========================================================
# LOAD CONFIG
# ==========================================================

def load_config():
    with open(CONFIG_PATH, "r") as f:
        return yaml.safe_load(f)


# ==========================================================
# LOAD MASTER DATAFRAME
# ==========================================================

def load_master_dataframe():

    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql("""
        SELECT *
        FROM financial_ratios
        WHERE year='Mar 2024'
    """, conn)

    companies = pd.read_sql("""
        SELECT *
        FROM companies
    """, conn)

    sectors = pd.read_sql("""
        SELECT *
        FROM sectors
    """, conn)

    market = pd.read_sql("""
        SELECT
            company_id,
            market_cap_crore,
            pe_ratio,
            pb_ratio,
            dividend_yield_pct
        FROM market_cap
        WHERE year=2024
    """, conn)

    # NEW: Load Profit & Loss data
    profit = pd.read_sql("""
        SELECT
            company_id,
            year,
            sales,
            net_profit,
            operating_profit
        FROM profitandloss
        WHERE year='Mar 2024'
    """, conn)

    conn.close()

    # Merge Financial Ratios + Companies
    df = ratios.merge(
        companies,
        left_on="company_id",
        right_on="id",
        how="left"
    )

    # Merge Sectors
    df = df.merge(
        sectors,
        on="company_id",
        how="left"
    )

    # Merge Market Cap
    df = df.merge(
        market,
        on="company_id",
        how="left"
    )

    # Merge Profit & Loss
    df = df.merge(
        profit,
        on=["company_id", "year"],
        how="left"
    )

    # Remove duplicate columns
    for col in ["id", "id_x", "id_y"]:
        if col in df.columns:
            df.drop(columns=col, inplace=True)

    return df

# ==========================================================
# NORMALIZE METRIC (P10/P90 WINSORIZATION)
# ==========================================================

def normalize_metric(series, inverse=False):
    """
    Normalize a metric to a 0-100 score using
    P10/P90 winsorization.

    inverse=True is used where lower values are better
    (like Debt to Equity).
    """

    s = pd.to_numeric(series, errors="coerce")

    # Remove missing values
    valid = s.dropna()

    if len(valid) == 0:
        return pd.Series(0, index=series.index)

    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)

    # Prevent divide-by-zero
    if p90 == p10:
        return pd.Series(50, index=series.index)

    # Winsorize
    clipped = s.clip(lower=p10, upper=p90)

    score = ((clipped - p10) / (p90 - p10)) * 100

    if inverse:
        score = 100 - score

    return score.fillna(0)


# ==========================================================
# APPLY FILTERS
# ==========================================================

def apply_filters(df, rules, preset_name):

    result = df.copy()

    # -------------------------
    # ROE
    # -------------------------

    if "roe_min" in rules:
        result = result[
            result["return_on_equity_pct"] >= rules["roe_min"]
        ]

    # -------------------------
    # Debt / Equity
    # Skip Financials
    # -------------------------

    if "debt_to_equity_max" in rules:

        # Debt-Free Bluechip should apply D/E to every company
        if preset_name == "debt_free_bluechip":

            result = result[
                result["debt_to_equity"] <=
                rules["debt_to_equity_max"]
            ]

        # All other screeners skip Financial companies
        else:

            financials = result[
                result["broad_sector"] == "Financials"
            ]

            others = result[
                result["broad_sector"] != "Financials"
            ]

            others = others[
                others["debt_to_equity"] <=
                rules["debt_to_equity_max"]
            ]

            result = pd.concat(
                [financials, others],
                ignore_index=True
            )

    # -------------------------
    # Free Cash Flow
    # -------------------------

    if "free_cash_flow_min" in rules:
        result = result[
            result["free_cash_flow_cr"] >= rules["free_cash_flow_min"]
        ]

    # -------------------------
    # Revenue CAGR
    # -------------------------

    if "revenue_cagr_5yr_min" in rules:
        result = result[
            result["revenue_cagr_5yr"] >= rules["revenue_cagr_5yr_min"]
        ]

    # -------------------------
    # PAT CAGR
    # -------------------------

    if "pat_cagr_5yr_min" in rules:
        result = result[
            result["pat_cagr_5yr"] >= rules["pat_cagr_5yr_min"]
        ]

    # -------------------------
    # EPS CAGR
    # -------------------------

    if "eps_cagr_5yr_min" in rules:
        result = result[
            result["eps_cagr_5yr"] >= rules["eps_cagr_5yr_min"]
        ]    

    # -------------------------
    # Operating Profit Margin
    # -------------------------

    if "operating_profit_margin_min" in rules:
        result = result[
            result["operating_profit_margin_pct"] >=
            rules["operating_profit_margin_min"]
        ]

    # -------------------------
    # PE
    # -------------------------

    if "pe_max" in rules:
        result = result[
            result["pe_ratio"] <= rules["pe_max"]
        ]

    # -------------------------
    # PB
    # -------------------------

    if "pb_max" in rules:
        result = result[
            result["pb_ratio"] <= rules["pb_max"]
        ]

    # -------------------------
    # Interest Coverage
    # Debt Free = Infinity
    # -------------------------

    if "interest_coverage_min" in rules:

        result["interest_coverage"] = result["interest_coverage"].replace(
            "Debt Free",
            float("inf")
        )

        result["interest_coverage"] = pd.to_numeric(
            result["interest_coverage"],
            errors="coerce"
        )

        result = result[
            result["interest_coverage"] >= rules["interest_coverage_min"]
        ]    

    # -------------------------
    # Dividend Yield
    # -------------------------

    if "dividend_yield_min" in rules:
        result = result[
            result["dividend_yield_pct"] >= rules["dividend_yield_min"]
        ]

    # -------------------------
    # Dividend Payout
    # -------------------------

    if "dividend_payout_ratio_max" in rules:
        result = result[
            result["dividend_payout_ratio_pct"] <=
            rules["dividend_payout_ratio_max"]
        ]

    # -------------------------
    # Market Cap
    # -------------------------

    if "market_cap_min" in rules:
        result = result[
            result["market_cap_crore"] >= rules["market_cap_min"]
        ]    

    # -------------------------
    # Net Profit
    # -------------------------

    if "net_profit_min" in rules:
        result = result[
            result["net_profit"] >= rules["net_profit_min"]
        ] 

    # -------------------------
    # Asset Turnover
    # -------------------------

    if "asset_turnover_min" in rules:
        result = result[
            result["asset_turnover"] >=
            rules["asset_turnover_min"]
        ]       

    # -------------------------
    # Sales
    # -------------------------

    if "sales_min" in rules:
        result = result[
            result["sales"] >= rules["sales_min"]
        ]

    result = result.sort_values(
        by="quality_score",
        ascending=False
    )

    return result



# ==========================================================
# CALCULATE COMPOSITE QUALITY SCORE
# ==========================================================

def calculate_composite_score(df):

    result = df.copy()

    # -------------------------
    # PROFITABILITY (35%)
    # -------------------------

    # Sector-relative ROE
    roe_score = (
        result.groupby("broad_sector")["return_on_equity_pct"]
        .transform(normalize_metric)
    )

    # Sector-relative ROCE
    roce_score = (
        result.groupby("broad_sector")["roce_percentage"]
        .transform(normalize_metric)
    )

    # Sector-relative Net Profit Margin
    npm_score = (
        result.groupby("broad_sector")["net_profit_margin_pct"]
        .transform(normalize_metric)
    )
    profitability = (
        roe_score * 0.15 +
        roce_score * 0.10 +
        npm_score * 0.10
    )

    result["profitability_score"] = profitability

    # -------------------------
    # CASH QUALITY (30%)
    # -------------------------

    # Sector-relative Free Cash Flow
    fcf_score = (
        result.groupby("broad_sector")["free_cash_flow_cr"]
        .transform(normalize_metric)
    )

    # Sector-relative Cash From Operations
    cfo_pat_score = (
        result.groupby("broad_sector")["cash_from_operations_cr"]
        .transform(normalize_metric)
    )

    fcf_positive = (
        result["free_cash_flow_cr"] > 0
    ).astype(int) * 100

    cash_quality = (
        fcf_score * 0.15 +
        cfo_pat_score * 0.10 +
        fcf_positive * 0.05
    )

    result["cash_quality_score"] = cash_quality

    # -------------------------
    # GROWTH SCORE
    # -------------------------

    # Sector-relative Revenue CAGR
    revenue_score = (
        result.groupby("broad_sector")["revenue_cagr_5yr"]
        .transform(normalize_metric)
    )

    # Sector-relative PAT CAGR
    pat_score = (
        result.groupby("broad_sector")["pat_cagr_5yr"]
        .transform(normalize_metric)
    )

    # Sector-relative EPS CAGR
    eps_score = (
        result.groupby("broad_sector")["eps_cagr_5yr"]
        .transform(normalize_metric)
    )

    growth_score = (
        revenue_score * 0.40 +
        pat_score * 0.40 +
        eps_score * 0.20
    )

    result["revenue_score"] = revenue_score
    result["pat_score"] = pat_score
    result["eps_score"] = eps_score

    result["growth_score"] = growth_score.round(2)

    # -------------------------
    # LEVERAGE (15%)
    # -------------------------

    # Default scores
    # Sector-relative Debt to Equity
    de_score = (
        result.groupby("broad_sector")["debt_to_equity"]
        .transform(lambda x: normalize_metric(x, inverse=True))
    )

    # Sector-relative Interest Coverage
    icr_score = (
        result.groupby("broad_sector")["interest_coverage"]
        .transform(normalize_metric)
    )

    # Financial companies: don't use D/E and Interest Coverage
    financial_mask = result["broad_sector"] == "Financials"

    de_score.loc[financial_mask] = 100
    icr_score.loc[financial_mask] = 100

    leverage = (
        de_score * 0.10 +
        icr_score * 0.05
    )

    result["leverage_score"] = leverage

    # -------------------------
    # FINAL SCORE
    # -------------------------


    # ======================================================
    # SECTOR-RELATIVE COMPOSITE SCORE
    # ======================================================

    result["raw_quality_score"] = (
        result["profitability_score"] * 0.35 +
        result["cash_quality_score"] * 0.30 +
        result["growth_score"] * 0.20 +
        result["leverage_score"] * 0.15
    ).round(2)

    result["quality_score"] = (
        result.groupby("broad_sector")["profitability_score"]
        .transform(lambda x: normalize_metric(x))
        * 0.35
        +
        result.groupby("broad_sector")["cash_quality_score"]
        .transform(lambda x: normalize_metric(x))
        * 0.30
        +
        result.groupby("broad_sector")["growth_score"]
        .transform(lambda x: normalize_metric(x))
        * 0.20
        +
        result.groupby("broad_sector")["leverage_score"]
        .transform(lambda x: normalize_metric(x))
        * 0.15
    ).round(2)

    result = result.sort_values(
        by="quality_score",
        ascending=False
    )

    return result 

# ==========================================================
# DISPLAY
# ==========================================================

def display(df):

    if len(df) == 0:
        print("No companies found.\n")
        return

    print(
        f"{'Rank':<5}"
        f"{'Company':<15}"
        f"{'Score':<8}"
        f"{'ROE':<10}"
        f"{'D/E':<8}"
        f"{'Revenue':<12}"
    )

    print("-" * 65)

    for rank, (_, row) in enumerate(df.head(10).iterrows(), start=1):

        print(
            f"{rank:<5}"
            f"{row['company_id']:<15}"
            f"{row['quality_score']:<8.2f}"
            f"{row['return_on_equity_pct']:<10.2f}"
            f"{row['debt_to_equity']:<8.2f}"
            f"{row['revenue_cagr_5yr']:<12}"
        )


# ==========================================================
# RUN ALL PRESETS
# ==========================================================

def run_all_screeners(df, config):

    preset_names = [
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_bluechip",
        "turnaround_watch"
    ]

    for preset in preset_names:

        print("\n")
        print("=" * 80)
        print(preset.upper().replace("_", " "))
        print("=" * 80)

        filtered = apply_filters(
            df,
            config[preset],
            preset
        )

        print(f"Companies Found : {len(filtered)}\n")

        display(filtered)



# ==========================================================
# MAIN
# ==========================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CONFIG_COLUMN_MAP = {
    "roe_min": "return_on_equity_pct",
    "debt_to_equity_max": "debt_to_equity",
    "free_cash_flow_min": "free_cash_flow_cr",
    "revenue_cagr_5yr_min": "revenue_cagr_5yr",
    "pat_cagr_5yr_min": "pat_cagr_5yr",
    "pe_max": "pe_ratio",
    "pb_max": "pb_ratio",
    "dividend_yield_min": "dividend_yield_pct",
    "dividend_payout_ratio_max": "dividend_payout_ratio_pct",
    "sales_min": "sales"
}

def export_screener_output(df, config):

    workbook = Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Header Style
    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Mapping: Sheet Name -> Config Key
    presets = {
        "Quality Compounder": "quality_compounder",
        "Value Pick": "value_pick",
        "Growth Accelerator": "growth_accelerator",
        "Dividend Champion": "dividend_champion",
        "Debt Free Blue Chip": "debt_free_bluechip",
        "Turnaround Watch": "turnaround_watch"
    }

    for sheet_name, preset in presets.items():

        sheet = workbook.create_sheet(title=sheet_name)

        # Freeze first row
        sheet.freeze_panes = "A2"

        # Apply screener
        filtered = apply_filters(
            df,
            config[preset],
            preset
        )

        # Write column headers
        sheet.append(list(filtered.columns))

        # -------------------------
        # Format Header Row
        # -------------------------
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # -------------------------
        # Write Data
        # -------------------------
        for row in filtered.itertuples(index=False):
            sheet.append(list(row))

        # -------------------------
        # Enable Filter
        # -------------------------
        sheet.auto_filter.ref = sheet.dimensions

        # -------------------------
        # Auto-fit Column Widths
        # -------------------------
        for column in sheet.columns:

            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:

                if cell.value is not None:

                    length = len(str(cell.value))

                    if length > max_length:
                        max_length = length

            sheet.column_dimensions[column_letter].width = max_length + 2

        # -------------------------
        # Colour Cells Based On Preset Thresholds
        # -------------------------
        rules = config[preset]

        for rule_name, threshold in rules.items():

            # Skip unknown rules
            if rule_name not in CONFIG_COLUMN_MAP:
                continue

            dataframe_column = CONFIG_COLUMN_MAP[rule_name]

            # Skip if dataframe doesn't have this column
            if dataframe_column not in filtered.columns:
                continue

            # Find Excel column number
            excel_column = (
                filtered.columns.get_loc(dataframe_column) + 1
            )

            # Colour every data cell
            for row in range(2, sheet.max_row + 1):

                cell = sheet.cell(
                    row=row,
                    column=excel_column
                )

                value = cell.value

                if value is None:
                    continue

                if rule_name.endswith("_min"):

                    if value >= threshold:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

                elif rule_name.endswith("_max"):

                    if value <= threshold:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

    workbook.save("output/screener_output.xlsx")

    print("Excel exported successfully.")

def main():

    print("=" * 60)
    print("Loading Screener Engine")
    print("=" * 60)

    config = load_config()

    df = load_master_dataframe()

    df = calculate_composite_score(df)

    # Print every column in the dataframe
    print(df.columns.tolist())

    print(df[["company_id", "quality_score"]].head(10))

    print(f"Loaded Companies : {len(df)}")

    run_all_screeners(df, config)

    export_screener_output(df, config)


if __name__ == "__main__":
    main()